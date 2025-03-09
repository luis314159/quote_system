import os
import io
import sys
import logging
import subprocess
import tempfile
from pathlib import Path
from django.conf import settings

# Configure logging
logger = logging.getLogger(__name__)

def excel_to_pdf_libreoffice(excel_data):
    """
    Convierte un archivo Excel directamente a PDF usando LibreOffice.
    Recomendado para entornos Linux.
    
    Args:
        excel_data (bytes): Contenido binario del archivo Excel
        
    Returns:
        bytes: Contenido binario del PDF generado
    """
    logger.info("Convirtiendo Excel a PDF usando LibreOffice")
    
    try:
        # Crear archivos temporales para entrada y salida
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_in:
            temp_in.write(excel_data)
            input_path = temp_in.name
            
        # Determinar la ruta de salida (mismo nombre pero extensión .pdf)
        output_path = os.path.splitext(input_path)[0] + '.pdf'
        
        # Comando para convertir usando LibreOffice en modo headless
        libreoffice_path = getattr(settings, 'LIBREOFFICE_PATH', 'libreoffice')
        cmd = [
            libreoffice_path,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', os.path.dirname(input_path),
            input_path
        ]
        
        logger.info(f"Ejecutando comando: {' '.join(cmd)}")
        
        # Ejecutar el comando
        process = subprocess.Popen(
            cmd, 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE
        )
        stdout, stderr = process.communicate()
        
        if process.returncode != 0:
            logger.error(f"Error durante la conversión: {stderr.decode()}")
            raise Exception(f"Error al convertir Excel a PDF: {stderr.decode()}")
            
        # Leer el archivo PDF generado
        if os.path.exists(output_path):
            with open(output_path, 'rb') as pdf_file:
                pdf_data = pdf_file.read()
                
            # Limpiar archivos temporales
            os.unlink(input_path)
            os.unlink(output_path)
            
            logger.info("Conversión de Excel a PDF completada exitosamente")
            return pdf_data
        else:
            raise Exception(f"No se generó el archivo PDF de salida en {output_path}")
    
    except Exception as e:
        logger.error(f"Error durante la conversión de Excel a PDF: {str(e)}")
        logger.error("Stacktrace:", exc_info=True)
        
        # Intento de limpieza en caso de error
        if 'input_path' in locals() and os.path.exists(input_path):
            os.unlink(input_path)
        if 'output_path' in locals() and os.path.exists(output_path):
            os.unlink(output_path)
            
        raise


def excel_to_pdf_reportlab(excel_data):
    """
    Convierte un archivo Excel a PDF usando ReportLab y openpyxl.
    Recomendado para entornos Windows donde WeasyPrint o LibreOffice son problemáticos.
    
    Args:
        excel_data (bytes): Contenido binario del archivo Excel
        
    Returns:
        bytes: Contenido binario del PDF generado
    """
    logger.info("Convirtiendo Excel a PDF usando ReportLab")
    
    try:
        import io
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        # Cargar el Excel
        wb = load_workbook(io.BytesIO(excel_data))
        ws = wb.active
        
        # Crear un buffer para el PDF
        buffer = io.BytesIO()
        
        # Crear el documento PDF
        doc = SimpleDocTemplate(buffer, pagesize=letter, 
                               rightMargin=36, leftMargin=36,
                               topMargin=36, bottomMargin=36)
        
        elements = []
        
        # Estilos para el texto
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=14,
            alignment=1,  # Centered
            spaceAfter=12
        )
        
        # Obtener información de cuota si está disponible
        quote_number = ws["J6"].value if ws["J6"].value else "N/A"
        project_name = ws["C6"].value if ws["C6"].value else "N/A"
        date = ws["J4"].value if ws["J4"].value else "N/A"
        valid_until = ws["J5"].value if ws["J5"].value else "N/A"
        
        # Agregar título e información
        elements.append(Paragraph(f"Quote #{quote_number}", title_style))
        elements.append(Paragraph(f"Project: {project_name}", styles['Normal']))
        elements.append(Paragraph(f"Date: {date}", styles['Normal']))
        elements.append(Paragraph(f"Valid Until: {valid_until}", styles['Normal']))
        elements.append(Spacer(1, 12))
        
        # Determinar el rango de datos (filas de la tabla)
        table_header_row = 11  # Basado en tu código de Excel
        table_start_row = 12
        max_row = 0
        
        # Buscar la última fila con datos (antes de la fila TOTAL)
        for row in range(table_start_row, 50):  # Buscar hasta un límite razonable
            if ws[f"B{row}"].value == "TOTAL":
                max_row = row
                break
            if ws[f"B{row}"].value is not None:
                max_row = row
        
        if max_row < table_start_row:
            max_row = table_start_row  # Al menos incluir una fila
        
        # Extraer encabezados
        headers = []
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            cell_value = ws[f"{col}{table_header_row}"].value
            headers.append(cell_value if cell_value else "")
        
        # Extraer los datos de la tabla
        data = [headers]  # La primera fila es el encabezado
        
        for row in range(table_start_row, max_row + 2):  # +2 para incluir la fila TOTAL
            row_data = []
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                cell_value = ws[f"{col}{row}"].value
                row_data.append(cell_value if cell_value else "")
            data.append(row_data)
        
        # Crear la tabla
        table = Table(data, repeatRows=1)
        
        # Estilo de la tabla
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),  # Fila TOTAL
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Fila TOTAL
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (5, -1), 'RIGHT'),  # Alinear a la derecha las columnas de números
            ('ALIGN', (8, 1), (10, -1), 'RIGHT'),  # Alinear a la derecha las columnas de números
        ])
        table.setStyle(table_style)
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Agregar términos y condiciones
        terms_row = 34  # Según tu código original
        elements.append(Paragraph("<b>Delivery:</b> 6 weeks or advise required", styles['Normal']))
        elements.append(Paragraph("<b>Payment Terms:</b> Net 15 after shipping", styles['Normal']))
        elements.append(Paragraph("<b>Incoterms:</b> Fob Grupo ARGA Plant at Chihuahua", styles['Normal']))
        elements.append(Paragraph("<b>Notes:</b> We can provide logistic service Door to Door including customs clearance. To be quoted based on needs and weight.", styles['Normal']))
        
        # Generar el PDF
        doc.build(elements)
        
        # Obtener el contenido del buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        logger.info("Conversión de Excel a PDF con ReportLab completada exitosamente")
        return pdf_data
        
    except Exception as e:
        logger.error(f"Error durante la conversión de Excel a PDF con ReportLab: {str(e)}")
        logger.error("Stacktrace:", exc_info=True)
        raise


def excel_to_pdf(excel_data):
    """
    Función principal que decide qué método usar según la plataforma.
    
    Args:
        excel_data (bytes): Contenido binario del archivo Excel
        
    Returns:
        bytes: Contenido binario del PDF generado
    """
    # Usar el método más apropiado según la plataforma
    if sys.platform.startswith('win'):
        logger.info("Detectado entorno Windows, usando ReportLab para conversión a PDF")
        try:
            return excel_to_pdf_reportlab(excel_data)
        except Exception as e:
            logger.error(f"Error usando método ReportLab: {str(e)}")
            logger.warning("Intentando con método alternativo (LibreOffice)")
            return excel_to_pdf_libreoffice(excel_data)
    else:
        logger.info("Detectado entorno Linux/Unix, usando LibreOffice para conversión a PDF")
        try:
            return excel_to_pdf_libreoffice(excel_data)
        except Exception as e:
            logger.error(f"Error usando método LibreOffice: {str(e)}")
            logger.warning("Intentando con método alternativo (ReportLab)")
            return excel_to_pdf_reportlab(excel_data)