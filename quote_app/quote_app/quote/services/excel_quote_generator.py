import os
import logging
import io
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from django.conf import settings

from ..models import MaterialDensity, CompanyMaterialPrice, CompanyFinishPrice

# Configure logging
logger = logging.getLogger(__name__)

class ExcelQuoteGenerator:
    def __init__(self, user, project_data, is_internal=False):
        self.user = user
        self.company = user.company
        self.project_data = project_data
        self.is_internal = is_internal
        
        # Definir referencias de celdas en un diccionario centralizado
        # Esto facilita ajustar las referencias si el template cambia
        self.cell_refs = {
            'date': 'J4',           # Desplazado una columna a la derecha
            'valid_until': 'J5',    # Desplazado una columna a la derecha
            'quote_number': 'J6',
            'revision': 'J7',
            'project_name': 'C6',
            'company_name': 'C7',   # Desplazado una columna a la derecha
            'contact_name': 'C8',   # Desplazado una columna a la derecha
            'contact_email': 'C9',  # Desplazado una columna a la derecha
            'table_header_row': 11,
            'table_start_row': 12,
            'terms_row': 34
        }
        
        # Cargar la plantilla Excel
        template_path = os.path.join(settings.STATIC_ROOT, 'templates', 'quote_template.xlsx')
        if not os.path.exists(template_path):
            logger.error(f"Template file not found at {template_path}")
            raise FileNotFoundError(f"Excel template not found: {template_path}")
        
        try:
            self.wb = load_workbook(template_path)
            self.ws = self.wb.active
            logger.info(f"Template loaded successfully: {template_path}")
            
            # Verificar si las celdas de referencia existen en el template
            self._validate_template()
        except Exception as e:
            logger.error(f"Error loading Excel template: {str(e)}")
            raise
        
        # Definir estilos
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
            'border': self.thin_border,
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        # Preparar datos de materiales y precios
        self._load_materials_and_prices()

    def _validate_template(self):
        """Valida que las celdas de referencia existan en el template"""
        # Obtener dimensiones de la hoja
        if not self.ws.dimensions:
            logger.warning("No se pudo determinar las dimensiones de la hoja")
            return
            
        logger.info(f"Sheet dimensions: {self.ws.dimensions}")
        
        # Verificar celdas combinadas
        if self.ws.merged_cells:
            merged_ranges = [str(merged_range) for merged_range in self.ws.merged_cells.ranges]
            logger.info(f"Merged cells found: {merged_ranges}")
        
        # Registrar información sobre las celdas clave
        for name, ref in self.cell_refs.items():
            if isinstance(ref, str) and not ref.isdigit():  # Solo verificar referencias de celda, no números de fila
                try:
                    value = self.ws[ref].value
                    logger.info(f"Reference '{name}' at {ref}: {value}")
                except (KeyError, AttributeError):
                    logger.warning(f"Reference '{name}' at {ref} not found in template")

    def _load_materials_and_prices(self):
        """Carga los datos de materiales y precios de la base de datos"""
        try:
            self.material_densities = {
                material.material_type: material.density 
                for material in MaterialDensity.objects.all()
            }
            
            self.material_prices = {
                cp.material.id: {
                    'price_per_lb': float(cp.price_per_lb),
                    'density': cp.material.density,
                    'name': cp.material.name or cp.material.get_material_type_display()
                }
                for cp in CompanyMaterialPrice.objects.filter(
                    company=self.company,
                    is_active=True
                ).select_related('material')
            }
            
            logger.info(f"Loaded {len(self.material_prices)} material prices for company {self.company.name}")
        except Exception as e:
            logger.error(f"Error loading material data: {str(e)}")
            self.material_densities = {}
            self.material_prices = {}
    
    def find_main_cell_in_merged_range(self, cell_ref):
        """
        Encuentra la celda principal (top-left) de un rango combinado que contiene la celda especificada.
        
        Args:
            cell_ref: Referencia de celda (ej: 'A1', 'B5')
            
        Returns:
            str: Referencia de la celda principal o None si no está en un rango combinado
        """
        # Extraer columna y fila de la referencia de celda
        try:
            col_letter = ''.join(filter(str.isalpha, cell_ref))
            row = int(''.join(filter(str.isdigit, cell_ref)))
            col = column_index_from_string(col_letter)
        except (ValueError, TypeError) as e:
            logger.error(f"Error parsing cell reference {cell_ref}: {str(e)}")
            return None
            
        # Buscar si la celda está en algún rango combinado
        for merged_range in self.ws.merged_cells.ranges:
            if (merged_range.min_col <= col <= merged_range.max_col and 
                merged_range.min_row <= row <= merged_range.max_row):
                # Calcular la celda principal (top-left) de la combinación
                min_col_letter = get_column_letter(merged_range.min_col)
                main_cell = f"{min_col_letter}{merged_range.min_row}"
                logger.debug(f"Cell {cell_ref} is in merged range {str(merged_range)}, main cell is {main_cell}")
                return main_cell
                
        # Si llegamos aquí, la celda no está en un rango combinado
        return None
    
    def safe_write_cell(self, cell_ref, value):
        """
        Escribe un valor en una celda de manera segura, manejando celdas combinadas.
        
        Args:
            cell_ref: Referencia de celda (ej: 'A1', 'B5')
            value: Valor a escribir en la celda
        """
        try:
            # Verificar si la celda está en un rango combinado
            main_cell = self.find_main_cell_in_merged_range(cell_ref)
            target_cell = main_cell if main_cell else cell_ref
            
            # Escribir el valor
            self.ws[target_cell] = value
            logger.debug(f"Successfully wrote value to cell {target_cell}")
            return True
        except Exception as e:
            logger.error(f"Error writing to cell {cell_ref}: {str(e)}")
            return False

    def _fill_header_info(self):
        """Rellena la información del encabezado en la plantilla Excel."""
        logger.info("Filling header information")
        
        # Número de cotización y fechas
        current_date = datetime.now()
        # Calcular fecha de validez (30 días después de la fecha actual)
        valid_date = current_date + timedelta(days=30)
        quote_number = f"AR{current_date.strftime('%d%m%Y%H%M')}"
        
        # Escribe las fechas y número de cotización
        self.safe_write_cell(self.cell_refs['date'], current_date.strftime('%m/%d/%Y'))
        self.safe_write_cell(self.cell_refs['valid_until'], valid_date.strftime('%m/%d/%Y'))
        self.safe_write_cell(self.cell_refs['quote_number'], quote_number)
        self.safe_write_cell(self.cell_refs['revision'], '1')
        
        # Información del cliente
        self.safe_write_cell(self.cell_refs['company_name'], self.company.name)
        self.safe_write_cell(self.cell_refs['contact_name'], self.company.contact_name)
        self.safe_write_cell(self.cell_refs['contact_email'], self.company.contact_email)
        
        # Proyecto
        if 'project_name' in self.project_data:
            self.safe_write_cell(self.cell_refs['project_name'], self.project_data['project_name'])
        
        # Aplicar formato a las cabeceras de la tabla
        self._format_table_headers()

    def _format_table_headers(self):
        """Aplica formato a las cabeceras de la tabla"""
        header_row = self.cell_refs['table_header_row']
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            cell_ref = f'{col}{header_row}'
            try:
                cell = self.ws[cell_ref]
                # Aplicar estilos solo si la celda existe y tiene un valor
                if cell.value:
                    cell.font = self.header_style['font']
                    cell.fill = self.header_style['fill']
                    cell.border = self.header_style['border']
                    cell.alignment = self.header_style['alignment']
                    logger.debug(f"Applied formatting to header cell {cell_ref}")
            except Exception as e:
                logger.warning(f"Failed to format header cell {cell_ref}: {str(e)}")

    def _add_components(self):
        """Añade los componentes a la tabla del Excel."""
        logger.info("Adding components to table")
        
        # Define la fila inicial donde insertar componentes
        start_row = self.cell_refs['table_start_row']
        
        # Calcular costes para cada componente
        costs_data, total_volume, total_weight, total_cost = self.calculate_component_costs()
        
        # Tanto para uso interno como para cliente, mostrar todos los componentes
        # pero con diferentes niveles de detalle
        self._add_detailed_components(costs_data, start_row)
        total_row = start_row + len(costs_data) + 1
        
        # Añadir fila de totales
        self._add_total_row(total_row, total_volume, total_weight, total_cost)

    def _add_detailed_components(self, costs_data, start_row):
        """Añade los componentes detallados"""
        for i, item in enumerate(costs_data):
            row = start_row + i
            
            # Información básica que se muestra tanto para interno como para cliente
            self.safe_write_cell(f'B{row}', item['component'])
            self.safe_write_cell(f'C{row}', item['material'])
            self.safe_write_cell(f'I{row}', item['quantity'])
            
            # Información detallada de costos solo para uso interno
            if self.is_internal:
                self.safe_write_cell(f'D{row}', f"{item['volume']:.2f} in³")
                self.safe_write_cell(f'E{row}', f"{item['weight']:.2f} lbs")
                self.safe_write_cell(f'F{row}', f"${item['price_per_pound']:.2f}/lb")
                self.safe_write_cell(f'J{row}', f"${item['unit_cost']:.2f}")
                self.safe_write_cell(f'K{row}', f"${item['subtotal']:.2f}")
            
            # Aplicar bordes a todas las celdas de la fila
            self._apply_borders_to_row(row)

    def _add_total_row(self, row, total_volume, total_weight, total_cost):
        """Añade la fila de totales"""
        self.safe_write_cell(f'B{row}', "TOTAL")
        
        # Añadir totales de volumen y peso en caso de vista interna
        if self.is_internal:
            self.safe_write_cell(f'D{row}', f"{total_volume:.2f} in³")
            self.safe_write_cell(f'E{row}', f"{total_weight:.2f} lbs")
            self.safe_write_cell(f'K{row}', f"${total_cost:.2f}")
        else:
            # Para el cliente, solo mostrar el total del proyecto
            self.safe_write_cell(f'K{row}', f"${total_cost:.2f}")
        
        # Aplicar formato a la fila de totales
        self._apply_total_row_format(row)

    def _apply_borders_to_row(self, row):
        """Aplica bordes a todas las celdas de una fila"""
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            try:
                cell = self.ws[f'{col}{row}']
                cell.border = self.thin_border
            except Exception as e:
                logger.debug(f"Failed to apply border to cell {col}{row}: {str(e)}")

    def _apply_total_row_format(self, row):
        """Aplica formato a la fila de totales"""
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            try:
                cell = self.ws[f'{col}{row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.border = self.thin_border
            except Exception as e:
                logger.debug(f"Failed to format total row cell {col}{row}: {str(e)}")

    def calculate_component_costs(self):
        """Calculate costs for all components considering material weight, quantity and price per pound."""
        logger.info("Calculating component costs")
        costs_data = []
        total_volume = 0
        total_weight = 0
        total_cost = 0

        # Verificar que los datos necesarios estén presentes
        required_keys = ['components', 'materials', 'quantities', 'volumes']
        for key in required_keys:
            if key not in self.project_data or not self.project_data[key]:
                logger.error(f"Missing required data: {key}")
                return [], 0, 0, 0
                
        # Verificar que las listas tengan la misma longitud
        list_lengths = [len(self.project_data[key]) for key in required_keys]
        if len(set(list_lengths)) > 1:
            logger.error(f"Mismatched data lengths: {list_lengths}")
            return [], 0, 0, 0

        for comp, mat_id, qty, vol in zip(
            self.project_data['components'],
            self.project_data['materials'],
            self.project_data['quantities'],
            self.project_data['volumes']
        ):
            try:
                # Convert values to appropriate types
                volume = float(vol)
                quantity = int(qty)
                mat_id = int(mat_id)
                
                if mat_id not in self.material_prices:
                    logger.warning(f"No price found for material ID {mat_id} in company {self.company.name}")
                    continue
                    
                material_data = self.material_prices[mat_id]
                
                # Calculate weight using material density
                weight = volume * material_data['density']  # Weight per unit in pounds
                
                # Calculate costs
                unit_cost = weight * material_data['price_per_lb']  # Cost for one unit
                subtotal = unit_cost * quantity  # Cost for all units of this component
                
                # Apply finish multiplier if specified
                finish_multiplier = self._get_finish_multiplier()
                subtotal *= finish_multiplier
                
                # Update totals
                total_volume += volume * quantity
                total_weight += weight * quantity
                total_cost += subtotal
                
                costs_data.append({
                    'component': comp,
                    'material': material_data['name'],
                    'quantity': quantity,
                    'volume': volume,
                    'weight': weight,
                    'unit_cost': unit_cost,
                    'subtotal': subtotal,
                    'price_per_pound': material_data['price_per_lb'],
                    'finish_multiplier': finish_multiplier
                })
                
                logger.debug(f"Processed component: {comp}, material: {material_data['name']}, cost: ${subtotal:.2f}")
                
            except (ValueError, TypeError) as e:
                logger.error(f"Error processing component {comp}: {str(e)}")
                continue
        
        logger.info(f"Total cost calculation: ${total_cost:.2f}")
        return costs_data, total_volume, total_weight, total_cost

    def _get_finish_multiplier(self):
        """Obtiene el multiplicador de acabado si está especificado"""
        finish_name = self.project_data.get('project_finish')
        finish_multiplier = 1.0
        
        if finish_name:
            try:
                finish_price = CompanyFinishPrice.objects.get(
                    company=self.company,
                    finish__name=finish_name,
                    is_active=True
                )
                finish_multiplier = float(finish_price.price_multiplier)
                logger.info(f"Applied finish multiplier {finish_multiplier} for {finish_name}")
            except CompanyFinishPrice.DoesNotExist:
                logger.warning(f"No finish price found for {finish_name} in company {self.company.name}")
        
        return finish_multiplier

    def _add_terms_and_delivery(self):
        """Añade información de términos y entrega en la parte inferior del Excel."""
        logger.info("Adding terms and delivery information")
        
        terms_row = self.cell_refs['terms_row']
        
        # Términos estándar (se pueden personalizar según necesidades)
        self.safe_write_cell(f'C{terms_row}', "6 weeks or advise required")  # Entrega
        self.safe_write_cell(f'C{terms_row+1}', "Net 15 after shipping")  # Términos de pago
        self.safe_write_cell(f'C{terms_row+2}', "Fob Grupo ARGA Plant at Chihuahua")  # Incoterms
        
        # Notas adicionales
        notes = "We can provide logistic service Door to Door including customs clerance. To be quoted base on needs and weight."
        self.safe_write_cell(f'C{terms_row+3}', notes)

    def generate_excel(self):
        """Genera el archivo Excel con todos los datos de la cotización."""
        logger.info("Generating Excel quote")
        
        try:
            # Paso 1: Rellenar la información del encabezado
            self._fill_header_info()
            
            # Paso 2: Añadir los componentes a la tabla
            self._add_components()
            
            # Paso 3: Añadir términos y condiciones
            self._add_terms_and_delivery()
            
            # Guardar el Excel en un buffer de memoria
            buffer = io.BytesIO()
            self.wb.save(buffer)
            buffer.seek(0)
            
            logger.info("Excel quote generated successfully")
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel quote: {str(e)}")
            logger.error("Stacktrace:", exc_info=True)
            raise